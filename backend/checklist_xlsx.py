"""Parse a downloaded Beckett (or similar) checklist .xlsx into structured card rows.

The workbook has one sheet per section (Base, Inserts, Autographs, Memorabilia,
Variations, ...). Each sheet starts with a section-title row and a "N cards" count
row, then rows of: card_number, "Player,", "Team", [tag like 'RC']. Insert/auto
sheets pack several sub-sections, each with its own title row — those titles are the
`subset`. We deterministically extract every card (no AI needed — the file is
already structured), skipping the redundant aggregate sheets.
"""
import base64
import io
import re

# Aggregate sheets that just re-list everything already on the granular sheets.
_SKIP_SHEETS = {"full checklist", "team sets", "team set"}
_COUNT_RE = re.compile(r"^\d+\s+cards?\b", re.I)
# A serial print run mentioned in a section title, e.g. "Gold /50" -> 50.
_NUMBERED_RE = re.compile(r"/\s*(\d[\d,]*)\b")


def _clean(v) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _numbered_from(text: str):
    m = _NUMBERED_RE.search(text or "")
    if not m:
        return None
    try:
        return int(m.group(1).replace(",", ""))
    except ValueError:
        return None


def _parse_sheet(ws, sheet_name: str) -> list:
    """Walk one worksheet, tracking the current section title as `subset`."""
    section = sheet_name
    out = []
    for row in ws.iter_rows(values_only=True):
        first = _clean(row[0])
        col1 = _clean(row[1]) if len(row) > 1 else ""
        col2 = _clean(row[2]) if len(row) > 2 else ""
        vals = [v for v in (first, col1, col2) if v]  # whitespace-only cells count as empty
        if not vals:
            continue
        # "200 cards" style count row — skip.
        if _COUNT_RE.match(first):
            continue
        # A lone non-empty value in column A is a section title (skip blank titles).
        if len(vals) == 1 and first and not col1 and not col2:
            section = first
            continue
        # Otherwise it's a card row: number, "Player,", team, [tag].
        player = col1.rstrip(",").strip()
        if not player:
            continue
        team = col2
        tag = _clean(row[3]) if len(row) > 3 else ""
        out.append({
            "card_number": first or None,
            "player": player,
            "team": team or None,
            "subset": section or None,
            "numbered_to": _numbered_from(section),
            "parallel": None,
            "rookie": bool(tag) and tag.upper().replace("-", "") in ("RC", "ROOKIE"),
        })
    return out


def parse_checklist_xlsx(data_base64: str) -> list:
    """Decode a base64 .xlsx and return a deduped list of card-row dicts."""
    import openpyxl
    raw = base64.b64decode(data_base64.split(",", 1)[-1])  # tolerate data: URLs
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows, seen = [], set()
    for name in wb.sheetnames:
        if name.strip().lower() in _SKIP_SHEETS:
            continue
        for r in _parse_sheet(wb[name], name):
            key = (r["subset"], r["card_number"], r["player"])
            if key in seen:
                continue
            seen.add(key)
            rows.append(r)
        if len(rows) >= 20000:  # sanity cap
            break
    wb.close()
    return rows
