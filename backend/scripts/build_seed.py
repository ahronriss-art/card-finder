"""One-time: convert the 'Final' sheet of Sports Card Shop.xlsx into shops_seed.json.

Maps columns by header name (robust to reordering), cleans values, dedupes on
(name, full_address), and writes backend/data/shops_seed.json which the app uses
to seed the database on first startup (works for both SQLite and Render Postgres).
"""
import json
import os
import openpyxl

XLSX = os.environ.get("SHOPS_XLSX", "/Users/ahronriss/Downloads/Sports Card Shop.xlsx")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "shops_seed.json")

# header text in the sheet  ->  field name in our model
HEADER_MAP = {
    "shop name": "name",
    "website": "website",
    "phone": "phone",
    "full_address": "full_address",
    "city": "city",
    "state": "state",
    "rating": "rating",
    "reviews": "reviews",
    "email": "email",
    "contact way": "contact_way",
    "tiktok handle": "tiktok",
    "whatnot handle": "whatnot",
    "direct account with topps/fanatics": "topps_fanatics",
    "buy from wholesalers": "buys_wholesale",
    "direct account with tcg": "tcg_account",
    "willing to wholesale with our wholesale department?": "willing_to_wholesale",
    "collectors they've been selling with": "collectors",
}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        # drop leftover spreadsheet formulas
        if v.startswith("="):
            return None
        return v or None
    return v


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Final"]
    rows = list(ws.iter_rows(values_only=True))
    header = [(h or "").strip().lower() if isinstance(h, str) else h for h in rows[0]]

    # build index -> field
    col_field = {}
    for idx, h in enumerate(header):
        if isinstance(h, str) and h in HEADER_MAP:
            col_field[idx] = HEADER_MAP[h]
    # column J (index 9) is an unnamed "contacted?" status flag (e.g. "yes(Mike)")
    col_field.setdefault(9, "contacted")

    shops = []
    seen = set()
    for r in rows[1:]:
        rec = {}
        for idx, field in col_field.items():
            if idx < len(r):
                rec[field] = clean(r[idx])
        name = rec.get("name")
        if not name:
            continue
        # numeric coercion
        for nf in ("rating", "reviews"):
            if isinstance(rec.get(nf), str):
                try:
                    rec[nf] = float(rec[nf])
                except ValueError:
                    rec[nf] = None
        if rec.get("reviews") is not None:
            rec["reviews"] = int(rec["reviews"])
        rec["shop_type"] = "shop"
        key = (name.lower(), (rec.get("full_address") or "").lower())
        if key in seen:
            continue
        seen.add(key)
        shops.append(rec)

    # --- Sheet1: larger raw list; add shops not already in Final ---
    if "Sheet1" in wb.sheetnames:
        s1 = list(wb["Sheet1"].iter_rows(values_only=True))[1:]  # name,site,type,phone,address,city,state,rating,reviews
        seen_names = {s["name"].lower() for s in shops}
        added = 0
        for r in s1:
            name = clean(r[0]) if r else None
            if not name:
                continue
            key = (name.lower(), clean(r[4]).lower() if len(r) > 4 and clean(r[4]) else "")
            if key in seen or name.lower() in seen_names:
                continue
            seen.add(key)
            seen_names.add(name.lower())
            rating = r[7] if len(r) > 7 else None
            reviews = r[8] if len(r) > 8 else None
            shops.append({
                "name": name, "website": clean(r[1]) if len(r) > 1 else None,
                "phone": clean(r[3]) if len(r) > 3 else None,
                "full_address": clean(r[4]) if len(r) > 4 else None,
                "city": clean(r[5]) if len(r) > 5 else None,
                "state": clean(r[6]) if len(r) > 6 else None,
                "rating": float(rating) if isinstance(rating, (int, float)) else None,
                "reviews": int(reviews) if isinstance(reviews, (int, float)) else None,
                "shop_type": "shop",
            })
            added += 1
        print(f"Added {added} extra shops from Sheet1")

    # --- Whatnot Breakers: online breakers, tagged separately ---
    wn_sheet = "Whatnot Breakers Card Shops"
    if wn_sheet in wb.sheetnames:
        wn = list(wb[wn_sheet].iter_rows(values_only=True))[1:]
        # cols: handle, person, instagram, phone, location, whatnot, ebay, extra, notes, additional
        wn_seen = set()
        added = 0
        for r in wn:
            handle = clean(r[0]) if r else None
            if not handle or handle.lower() in wn_seen:
                continue
            wn_seen.add(handle.lower())
            person = clean(r[1]) if len(r) > 1 else None
            ebay = clean(r[6]) if len(r) > 6 else None
            extras = [clean(r[i]) for i in (7, 8, 9) if len(r) > i and clean(r[i])]
            note_bits = []
            if person:
                note_bits.append(f"Contact: {person}")
            if ebay:
                note_bits.append(f"eBay: {ebay}")
            note_bits += extras
            shops.append({
                "name": handle,
                "instagram": clean(r[2]) if len(r) > 2 else None,
                "phone": clean(r[3]) if len(r) > 3 else None,
                "full_address": clean(r[4]) if len(r) > 4 else None,
                "whatnot": clean(r[5]) if len(r) > 5 else handle,
                "notes": " | ".join(note_bits) or None,
                "shop_type": "whatnot_breaker",
            })
            added += 1
        print(f"Added {added} Whatnot breakers")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(shops, f, indent=1, ensure_ascii=False)
    print(f"Wrote {len(shops)} shops to {os.path.abspath(OUT)}")
    # quick stats
    states = {}
    for s in shops:
        states[s.get("state")] = states.get(s.get("state"), 0) + 1
    print("Top states:", sorted(states.items(), key=lambda x: -x[1])[:8])


if __name__ == "__main__":
    main()
