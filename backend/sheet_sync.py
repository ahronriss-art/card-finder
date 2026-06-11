"""One-way sync from the live Google Sheet into card_shops.

Downloads the whole workbook (xlsx export — link-viewable, no Google auth)
and syncs all three tabs:
  - "Final"                       -> physical shops (rich columns)
  - "Sheet1"                      -> physical shops (basic columns)
  - "Whatnot Breakers Card Shops" -> online breakers (shop_type=whatnot_breaker)

UPSERT by (name, full_address): existing shops get any non-empty sheet cells
applied (never blanked), new rows are inserted. `notes`/`update_log` are
website-owned — the sheet only sets notes on brand-new inserts, never on update.
"""
import io
import os
import httpx
import openpyxl
from sqlalchemy import select

from database import CardShop

SHEET_ID = os.getenv("SHEET_ID", "1t6oyf3VWtOBFxfFk-dB8G7zFjtPOcY1om4NkosDYmgE")
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# fields the sheet may overwrite on an existing shop (notes/update_log excluded)
WEBSITE_OWNED = {"notes", "update_log", "id", "created_at", "updated_at"}

FINAL_MAP = {
    "shop name": "name", "website": "website", "phone": "phone",
    "full_address": "full_address", "city": "city", "state": "state",
    "rating": "rating", "reviews": "reviews", "email": "email",
    "contact way": "contact_way", "tiktok handle": "tiktok", "whatnot handle": "whatnot",
    "direct account with topps/fanatics": "topps_fanatics",
    "buy from wholesalers": "buys_wholesale", "direct account with tcg": "tcg_account",
    "willing to wholesale with our wholesale department?": "willing_to_wholesale",
    "collectors they've been selling with": "collectors",
}
SHEET1_MAP = {
    "name": "name", "site": "website", "phone": "phone", "full_address": "full_address",
    "city": "city", "state": "state", "rating": "rating", "reviews": "reviews",
    "email": "email", "status": "contacted",
}


def _clean(v):
    if v is None:
        return None
    v = str(v).strip()
    if not v or v.startswith("="):
        return None
    return v


def _num(rec, field, integer=False):
    if rec.get(field) is not None:
        try:
            rec[field] = int(float(rec[field])) if integer else float(rec[field])
        except (ValueError, TypeError):
            rec[field] = None


def _rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    return rows[0] if rows else [], rows[1:] if rows else []


def _parse_final(ws):
    header, data = _rows(ws)
    hmap = {i: FINAL_MAP[(h or "").strip().lower()] for i, h in enumerate(header)
            if isinstance(h, str) and (h or "").strip().lower() in FINAL_MAP}
    hmap.setdefault(9, "contacted")  # unnamed status column; "Instagram Links" is a formula -> skipped
    out = []
    for r in data:
        rec = {f: _clean(r[i]) for i, f in hmap.items() if i < len(r)}
        if not rec.get("name"):
            continue
        _num(rec, "rating"); _num(rec, "reviews", integer=True)
        rec["shop_type"] = "shop"
        out.append(rec)
    return out


def _parse_sheet1(ws):
    header, data = _rows(ws)
    hmap = {i: SHEET1_MAP[(h or "").strip().lower()] for i, h in enumerate(header)
            if isinstance(h, str) and (h or "").strip().lower() in SHEET1_MAP}
    out = []
    for r in data:
        rec = {f: _clean(r[i]) for i, f in hmap.items() if i < len(r)}
        if not rec.get("name"):
            continue
        _num(rec, "rating"); _num(rec, "reviews", integer=True)
        rec["shop_type"] = "shop"
        out.append(rec)
    return out


def _parse_whatnot(ws):
    _, data = _rows(ws)
    out = []
    for r in data:
        handle = _clean(r[0]) if r else None
        if not handle:
            continue
        person = _clean(r[1]) if len(r) > 1 else None
        ebay = _clean(r[6]) if len(r) > 6 else None
        extras = [_clean(r[i]) for i in (7, 8, 9) if len(r) > i and _clean(r[i])]
        note_bits = ([f"Contact: {person}"] if person else []) + ([f"eBay: {ebay}"] if ebay else []) + extras
        out.append({
            "name": handle,
            "instagram": _clean(r[2]) if len(r) > 2 else None,
            "phone": _clean(r[3]) if len(r) > 3 else None,
            "full_address": _clean(r[4]) if len(r) > 4 else None,
            "whatnot": (_clean(r[5]) if len(r) > 5 else None) or handle,
            "notes": " | ".join(note_bits) or None,
            "shop_type": "whatnot_breaker",
        })
    return out


def _key(name, addr):
    return (name or "").lower().strip() + "|" + (addr or "").lower().strip()


async def sync_from_sheet(session) -> dict:
    """Fetch the workbook and upsert all tabs. Returns a summary dict."""
    resp = httpx.get(XLSX_URL, timeout=45, follow_redirects=True)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    records = []
    if "Final" in wb.sheetnames:
        records += _parse_final(wb["Final"])
    if "Sheet1" in wb.sheetnames:
        records += _parse_sheet1(wb["Sheet1"])
    if "Whatnot Breakers Card Shops" in wb.sheetnames:
        records += _parse_whatnot(wb["Whatnot Breakers Card Shops"])
    if not records:
        return {"checked": 0, "added": 0, "updated": 0, "fields_changed": 0}

    valid = {c.name for c in CardShop.__table__.columns}
    result = await session.execute(select(CardShop))
    by_key = {}
    for s in result.scalars().all():
        by_key.setdefault(_key(s.name, s.full_address), s)

    added = updated = fields_changed = 0
    for rec in records:
        key = _key(rec.get("name"), rec.get("full_address"))
        existing = by_key.get(key)
        if existing is None:
            data = {k: v for k, v in rec.items() if k in valid and v is not None}
            new_shop = CardShop(**data)
            session.add(new_shop)
            by_key[key] = new_shop
            added += 1
            continue
        changed_here = False
        for field, value in rec.items():
            if field not in valid or field in WEBSITE_OWNED or value is None or value == "":
                continue
            if getattr(existing, field, None) != value:
                setattr(existing, field, value)
                fields_changed += 1
                changed_here = True
        if changed_here:
            updated += 1

    await session.commit()
    return {"checked": len(records), "added": added, "updated": updated, "fields_changed": fields_changed}
